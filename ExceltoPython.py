import openpyxl as xl

wb = xl.load_workbook('example.xlsx')

sn = wb.sheetnames

print(sn)

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(sheet1)
print(cellA1)

print(cellA1.value)  # value method gives us the data in the cell
print(type(cellA1.value))  # recognizes the data type

print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

print(sheet1.cell(2, 2).value)  # rows and columns start at 1

print(sheet1.max_row)  # how many rows
print(sheet1.max_column)  # how many columns


for i in range(1, sheet1.max_row+1):
    print(sheet1.cell(i, 2).value)

# convert letters to numbers
print(xl.utils.get_column_letter(1))
print(xl.utils.get_column_letter(900))

print(xl.utils.column_index_from_string('AHP'))

for currentrow in sheet1['A1':'C3']:
    print(currentrow)
    for currentcell in currentrow:
        print(currentcell.coordinate, currentcell.value)

for currentrow in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, max_col=sheet1.max_column):
    print(currentrow)
    print(currentrow[0].value)
    print(currentrow[1].value)
    print(currentrow[2].value)
