from twilio.rest import Client
import keys
import random
from urllib.request import urlopen
from bs4 import BeautifulSoup
from urllib.request import urlopen, Request
import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.styles import numbers
import locale


webpage = 'https://www.coingecko.com/en/crypto-gainers-losers'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(webpage, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

print(soup.title.text)

table_rows = soup.findAll("tr")
# print(len(table_rows))
# print(table_rows[2:6])
# element = soup.find('td', class_="filter-item")
# print(element)


# putting webscraped data into excel

wb = xl.Workbook()

ws = wb.active

ws.title = 'Crypto Tracker'

ws['A1'] = 'Symbol'
ws['A1'].font = Font(name='Times New Roman', size=12, italic=False, bold=True)
ws.column_dimensions['A'].width = 10

ws['B1'] = 'Name'
ws['B1'].font = Font(name='Times New Roman', size=12, italic=False, bold=True)
ws.column_dimensions['B'].width = 10

ws['C1'] = 'Current Price'
ws['C1'].font = Font(name='Times New Roman', size=12, italic=False, bold=True)
ws.column_dimensions['C'].width = 20

ws['D1'] = 'Percent Change'
ws['D1'].font = Font(name='Times New Roman', size=12,
                     italic=False, bold=True)

ws.column_dimensions['D'].width = 20

ws['E1'] = "Yesterday's Price"
ws['E1'].font = Font(name='Times New Roman', size=12, italic=False, bold=True)
ws.column_dimensions['E'].width = 20

write_sheet = wb['Crypto Tracker']


i = 2
for row in table_rows[1:6]:
    td = list(row.findAll("td"))
    symbol_and_name_list = td[2].text.split()
    name = symbol_and_name_list[0]
    symbol = symbol_and_name_list[1]
    current_price = td[3].text
    percent_change = td[5].text

    # BREAK

    current_price_without_dollarsign = current_price.replace('$', '')
    round_current_price = round(float(current_price_without_dollarsign), 4)
    percent_change_without_sign = percent_change.replace('%', '')
    number_as_percent = float(percent_change_without_sign) / 100
    yesterday_price = float(
        current_price_without_dollarsign) / (1 + number_as_percent)
    round_yesterday_price = round(yesterday_price, 4)

    write_sheet.cell(i, 1).value = symbol
    write_sheet.cell(i, 2).value = name
    write_sheet.cell(i, 3).value = '$' + str(round_current_price)
    font = Font(color="008000")
    write_sheet.cell(i, 4).value = percent_change
    write_sheet.cell(i, 4).font = font
    # percent_change.font = font
    write_sheet.cell(i, 5).value = '$' + str(round_yesterday_price)
    i += 1

    for row in ws.iter_rows(min_row=2, max_row=6, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


wb.save('CryptoReport.xlsx')


# New website to send me a text about bitcoina and ethereum
webpage2 = 'https://www.coingecko.com/'

headers2 = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(webpage2, headers=headers2)

webpage2 = urlopen(req).read()

soup = BeautifulSoup(webpage2, 'html.parser')

print(soup.title.text)

table_rows = soup.findAll("tr")


i = 2
for row in table_rows[1:6]:
    td = list(row.findAll("td"))
    symbol_and_name_list = td[2].text.split()
    name = symbol_and_name_list[0]
    current_price = td[3].text.replace('\n', '')
    current_price_without_dollarsign = current_price.replace(
        '$', '').replace(',', '')
    percent_change = td[5].text.replace('\n', '')
    percent_change_without_sign = percent_change.replace('%', '')

    number_as_percent = float(percent_change_without_sign) / 100
    yesterday_price = float(
        current_price_without_dollarsign) / (1 + number_as_percent)

    locale.setlocale(locale.LC_ALL, '')  # set the locale to the user's default
    currency_price = locale.currency(yesterday_price, grouping=True)

    if name == 'Bitcoin':
        if abs(float(current_price_without_dollarsign) - float(yesterday_price)) > 5:
            client = Client(keys.account_sid, keys.auth_token)
            message = name + ' has changed to a price of ' + \
                str(current_price) + ' from ' + currency_price
            TWnumber = '+12147616829'
            myphone = '+12817506263'
            textmsg = client.messages.create(
                to=myphone, from_=TWnumber, body=message)
            print(textmsg.status)

    if name == 'Ethereum':
        if abs(float(current_price_without_dollarsign) - float(yesterday_price)) > 5:
            client = Client(keys.account_sid, keys.auth_token)
            message = name + ' has changed to a price of ' + \
                str(current_price) + ' from ' + currency_price
            TWnumber = '+12147616829'
            myphone = '+12817506263'
            textmsg = client.messages.create(
                to=myphone, from_=TWnumber, body=message)
            print(textmsg.status)
